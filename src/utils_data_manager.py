from datetime import datetime
import os
import pandas as pd
import logging

EXCEL_TAIL_ROWS = int(os.environ.get("QMANAGER_EXCEL_TAIL_ROWS", "2000"))

SHEETS_COLUMN_INDICES = {
    'date': 1,
    'vessel': 3,
    'voyage': 4,
    'bill': 6,
    'container': 13,
    'company': 17
}

class DataManager:
    EXCEL_COLUMN_MAPPINGS = {
        'container': ['Номер конт / тс'],
        'order': ['Номер заказа (заказ)'],
        'vessel': ['Судно / номер ТС (поставка)'],
        'date': ['Факт дата прибытия порт/свх (поставка)'],
        'bill': ['Коносамент / CMR (поставка)']
    }
    EXCEL_COLUMN_MAPPINGS_REPORT = {
        'container': ['Контейнер'],
        'order': ['Номер Заказа'],
        'vessel': ['Судно'],
        'date': ['Дата прибытия'],
        'bill': ['Коносамент']
    }

    def __init__(self, mode="logos", merge_mode: str = "order"):
        self.mode = mode
        self.merge_mode = (str(merge_mode).lower() if merge_mode else "order")
        self.latest_container_data = {}
        self.containers_by_unit = {}
        self.logger = logging.getLogger(f'DataManager.{mode}')

    def _find_column_index(self, columns, possible_names):
        for name in possible_names:
            matches = [i for i, col in enumerate(columns) if str(col).strip().lower() == name.lower()]
            if matches:
                return matches[0]
        self.logger.warning(f"Столбец не найден для имен: {possible_names}")
        return None

    def _parse_date(self, date_str):
        formats = [
            "%d.%m.%Y",
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%d.%m.%Y %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%d.%m.%Y %H:%M:%S"
        ]
        for fmt in formats:
            try:
                return datetime.strptime(str(date_str).strip(), fmt)
            except ValueError:
                continue
        return None

    def _extract_container_number(self, container_str):
        if not container_str or pd.isna(container_str):
            return None
        digits = ''.join(c for c in str(container_str) if c.isdigit())
        if len(digits) >= 7:
            return digits[-7:]
        return None

    def _process_excel_row(self, row, column_indices, filter_mode, filter_date_from, filter_date_to):
        date_str = str(row.iloc[column_indices['date']])
        parsed_date = self._parse_date(date_str)
        
        if not parsed_date:
            return None
        
        if filter_mode == 'period':
            parsed_date_normalized = parsed_date.replace(hour=0, minute=0, second=0, microsecond=0)
            if filter_date_from:
                filter_date_from_normalized = filter_date_from.replace(hour=0, minute=0, second=0, microsecond=0)
                if parsed_date_normalized < filter_date_from_normalized:
                    return None
            if filter_date_to:
                filter_date_to_normalized = filter_date_to.replace(hour=0, minute=0, second=0, microsecond=0)
                if parsed_date_normalized > filter_date_to_normalized:
                    return None
        else:
            days_ago = (datetime.now() - parsed_date).days
            if days_ago > 60:
                return 'STOP'
        
        container_str = str(row.iloc[column_indices['container']])
        container_number = self._extract_container_number(container_str)
        
        if not container_number:
            return None
        
        order = str(row.iloc[column_indices['order']]) if column_indices['order'] is not None else None
        vessel = str(row.iloc[column_indices['vessel']]) if column_indices['vessel'] is not None else None
        date = str(row.iloc[column_indices['date']]) if column_indices['date'] is not None else None
        bill = str(row.iloc[column_indices['bill']]) if column_indices['bill'] is not None else None
        if bill:
            bill = bill.replace(" ", "")
        
        return {
            'container': container_str,
            'container_number': container_number,
            'order': order,
            'vessel': vessel,
            'date': date,
            'bill': bill
        }

    def load_excel_data(self, excel_path, filter_mode='unlimited', filter_date_from=None, filter_date_to=None):
        self.logger.info(f"Начало загрузки Excel файла: {excel_path}")
        try:
            with pd.ExcelFile(excel_path) as excel_file:
                if filter_mode == 'period' or EXCEL_TAIL_ROWS <= 0:
                    df = pd.read_excel(excel_file)
                else:
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
                        ws = wb[wb.sheetnames[0]]
                        total_rows_in_sheet = int(ws.max_row or 0)
                        wb.close()

                        first_row_to_keep = max(2, total_rows_in_sheet - EXCEL_TAIL_ROWS + 1)
                        if first_row_to_keep > 2:
                            self.logger.info(
                                f"Excel optimization: читаем последние {EXCEL_TAIL_ROWS} строк "
                                f"(из ~{total_rows_in_sheet})"
                            )
                            df = pd.read_excel(excel_file, skiprows=range(1, first_row_to_keep))
                        else:
                            df = pd.read_excel(excel_file)
                    except Exception as opt_e:
                        self.logger.warning(f"Не удалось применить оптимизацию хвоста Excel, читаем весь файл: {opt_e}")
                        df = pd.read_excel(excel_file)
                
                self.logger.info(f"Excel прочитан: {len(df)} строк × {len(df.columns)} столбцов")
                
                columns = df.columns.tolist()
                column_indices = {}
                mappings = self.EXCEL_COLUMN_MAPPINGS_REPORT if self.mode == "report" else self.EXCEL_COLUMN_MAPPINGS
                
                for key, possible_names in mappings.items():
                    idx = self._find_column_index(columns, possible_names)
                    if idx is None:
                        raise ValueError(f"Не найден столбец для {key}. Возможные имена: {possible_names}")
                    column_indices[key] = idx
                
                valid_rows = 0
                valid_containers = 0
                start_row = None
                end_row = None
                
                for i in range(len(df) - 1, -1, -1):
                    row = df.iloc[i]
                    result = self._process_excel_row(row, column_indices, filter_mode, filter_date_from, filter_date_to)
                    
                    if result == 'STOP':
                        break
                    if result is None:
                        continue
                    
                    container_number = result['container_number']
                    if container_number not in self.containers_by_unit:
                        self.containers_by_unit[container_number] = []
                    
                    self.containers_by_unit[container_number].append({
                        'container': result['container'],
                        'order': result['order'],
                        'vessel': result['vessel'],
                        'date': result['date'],
                        'bill': result['bill']
                    })
                    
                    valid_rows += 1
                    valid_containers += 1
                    
                    if start_row is None:
                        start_row = i + 1
                    end_row = i + 1
                
                self.latest_container_data.clear()
                
                for suffix, containers in self.containers_by_unit.items():
                    latest_container = containers[0]
                    self.latest_container_data[latest_container['container']] = {
                        'company': 'GRAND-TRADE',
                        'order': latest_container['order'],
                        'vessel': latest_container['vessel'],
                        'date': latest_container['date'],
                        'bill': latest_container['bill']
                    }
                
                self.logger.info(f"Данные сохранены: {len(self.latest_container_data)} уникальных контейнеров")
                
                return {
                    'total_rows': len(df),
                    'valid_rows_range': (start_row, end_row),
                    'valid_containers': valid_containers
                }
                    
        except Exception as e:
            self.logger.error(f"Ошибка при загрузке Excel: {e}", exc_info=True)
            raise

    def load_sheets_data(self, credentials_file=None, spreadsheet_id=None, range_name=None, filter_mode='unlimited', filter_date_from=None, filter_date_to=None, log_callback=None):
        self.logger.info("Начало загрузки данных из Google Sheets")
        try:
            from src.utils_sheets_manager import GoogleSheetsManager
            
            if log_callback:
                log_callback("Подключение к Google Sheets API...")
            manager = GoogleSheetsManager(credentials_file, spreadsheet_id)
            if log_callback:
                log_callback("Запрос данных с сервера...")
            data = manager.get_data(range_name)
            
            self.logger.info(f"Получено {len(data)} строк из Google Sheets")
            
            self.latest_container_data.clear()
            
            valid_containers = 0
            
            for i in range(len(data) - 1, -1, -1):
                row = data[i]
                row_idx = len(data) - i + 1
                try:
                    if len(row) <= max(SHEETS_COLUMN_INDICES.values()):
                        continue
                    
                    container = str(row[SHEETS_COLUMN_INDICES['container']]).strip() if len(row) > SHEETS_COLUMN_INDICES['container'] else ""
                    if not container:
                        continue
                    
                    if container in self.latest_container_data:
                        continue
                    
                    date = str(row[SHEETS_COLUMN_INDICES['date']]).strip() if len(row) > SHEETS_COLUMN_INDICES['date'] else ""
                    vessel = str(row[SHEETS_COLUMN_INDICES['vessel']]).strip() if len(row) > SHEETS_COLUMN_INDICES['vessel'] else ""
                    voyage = str(row[SHEETS_COLUMN_INDICES['voyage']]).strip() if len(row) > SHEETS_COLUMN_INDICES['voyage'] else ""
                    bill = str(row[SHEETS_COLUMN_INDICES['bill']]).strip() if len(row) > SHEETS_COLUMN_INDICES['bill'] else ""
                    company = str(row[SHEETS_COLUMN_INDICES['company']]).strip() if len(row) > SHEETS_COLUMN_INDICES['company'] else ""
                    
                    if date:
                        parsed_date = self._parse_date(date)
                        if parsed_date:
                            if filter_mode == 'period':
                                parsed_date_normalized = parsed_date.replace(hour=0, minute=0, second=0, microsecond=0)
                                if filter_date_from:
                                    filter_date_from_normalized = filter_date_from.replace(hour=0, minute=0, second=0, microsecond=0)
                                    if parsed_date_normalized < filter_date_from_normalized:
                                        continue
                                if filter_date_to:
                                    filter_date_to_normalized = filter_date_to.replace(hour=0, minute=0, second=0, microsecond=0)
                                    if parsed_date_normalized > filter_date_to_normalized:
                                        continue
                            else:
                                days_ago = (datetime.now() - parsed_date).days
                                if days_ago > 60:
                                    break
                    
                    if bill:
                        bill = bill.replace(" ", "")
                    
                    self.latest_container_data[container] = {
                        'company': company if company else 'OTHER',
                        'order': None,
                        'vessel': vessel if vessel else '',
                        'date': date if date else '',
                        'bill': bill if bill else '',
                        'voyage': voyage if voyage else ''
                    }
                    
                    valid_containers += 1
                    
                except Exception as e:
                    self.logger.error(f"Ошибка при обработке строки {row_idx} Google Sheets: {e}")
                    continue
            
            self.logger.info(f"Данные из Google Sheets загружены: {valid_containers} контейнеров")
            
            return {
                'total_rows': len(data),
                'valid_containers': valid_containers
            }
            
        except Exception as e:
            self.logger.error(f"Ошибка при загрузке Google Sheets: {e}", exc_info=True)
            raise

    def process_data(self):
        self.logger.info("Начало организации контейнеров по юнитам")
        self.containers_by_unit = {}
        
        for container, row in self.latest_container_data.items():
            try:
                company = row.get("company", "")
                
                if self.mode == "sheets":
                    unit = row.get("bill") or ""
                else:
                    if self.merge_mode == "bl":
                        unit = row.get("bill")
                    elif company == "GRAND-TRADE":
                        order = row.get("order") or ""
                        vessel = row.get("vessel") or ""
                        unit = f"{order}|{vessel}" if order and vessel else order
                    else:
                        unit = row.get("bill")
                
                if not unit:
                    continue
                
                if unit not in self.containers_by_unit:
                    self.containers_by_unit[unit] = []
                
                if container not in self.containers_by_unit[unit]:
                    self.containers_by_unit[unit].append(container)
                    
            except Exception as e:
                self.logger.error(f"Ошибка при обработке контейнера '{container}': {e}")
                continue
        
        self.logger.info(f"Организация завершена: {len(self.containers_by_unit)} units")

    def get_container_data(self, container):
        return self.latest_container_data.get(container)

    def get_containers_by_unit(self, unit):
        result = self.containers_by_unit.get(unit)
        if not result:
            norm = (str(unit) if unit is not None else "").replace(" ", "")
            if self.merge_mode == "bl":
                result = [c for c, row in self.latest_container_data.items()
                          if (str(row.get('bill') or '').replace(' ', '')) == norm]
            else:
                if '|' in norm:
                    parts = norm.split('|', 1)
                    order_part = parts[0] if len(parts) > 0 else ""
                    vessel_part = parts[1] if len(parts) > 1 else ""
                    result = [c for c, row in self.latest_container_data.items()
                              if str(row.get('order') or '') == order_part and str(row.get('vessel') or '') == vessel_part]
                else:
                    result = [c for c, row in self.latest_container_data.items()
                              if str(row.get('order') or '') == norm]
        return result
